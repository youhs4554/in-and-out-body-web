from collections import defaultdict
import imaplib
import email
from email.policy import default
from datetime import datetime, timedelta, timezone

import base64
from io import BytesIO
import re
from PIL import Image
import boto3
from django.conf import settings
import requests
from .models import BodyResult, CodeInfo
from django.db.models import Q
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
import pandas as pd
import io as excel_io

# 전역 변수 S3 클라이언트 생성
s3_client = None


# boto3 반환 함수
def get_s3_client():
    global s3_client
    if s3_client is None:
        s3_client = boto3.client(
            's3',
            aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
            aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
            region_name=settings.AWS_S3_REGION_NAME
        )
    return s3_client


def generate_file_key(*args):  # ('front' + created_dt.png)
    return '-'.join(*args)


def verify_image(byte_string):
    """이미지 검증 함수"""
    try:
        # byte string을 이미지로 변환
        image_data = base64.b64decode(byte_string)
        image = Image.open(BytesIO(image_data))
        image.verify()  # 이미지 파일 검증
        return image_data
    except Exception as e:
        raise ValueError("Image verification failed") from e


def upload_image_to_s3(image_data, file_keys):
    """검증된 이미지를 S3에 업로드하는 함수"""
    file_name = generate_file_key(file_keys) + '.png'

    # AWS S3 클라이언트 생성
    s3 = get_s3_client()

    try:
        image = Image.open(BytesIO(image_data))  # 검증된 이미지 데이터로 이미지 객체 생성

        # 이미지를 BytesIO에 저장
        buffer = BytesIO()
        image.save(buffer, format='PNG')
        buffer.seek(0)
    except Exception as e:  # 이미지 처리 중 실패
        raise ValueError("Image processing failed, please check the image file format.") from e

    """시간이 많이 걸린다면, 비동기 처리를 고려해야 함"""

    # S3 버킷에 이미지 업로드
    try:
        s3.put_object(
            Bucket=settings.AWS_STORAGE_BUCKET_NAME,
            Key=file_name,
            Body=buffer,
            ContentType='image/png'
        )
    except Exception as e:  # AWS S3 이미지 업로드 실패
        raise Exception("Failed to upload image to S3") from e


def generate_presigned_url(file_keys, expiration=settings.AWS_PRESIGNED_EXPIRATION):
    file_name = generate_file_key(file_keys) + '.png'

    # AWS S3 클라이언트 생성
    s3 = get_s3_client()

    # presigned URL 생성
    try:
        presigned_url = s3.generate_presigned_url(
            'get_object',
            Params={
                'Bucket': settings.AWS_STORAGE_BUCKET_NAME,
                'Key': file_name
            },
            ExpiresIn=expiration  # URL 만료 시간 (초 단위)
        )
        return presigned_url
    except Exception as e:
        print(f"Error generating presigned URL: {e}")
        return None


def parse_userinfo(userinfo_obj):
    return {
        'user_id': userinfo_obj.id,
        'user_type': userinfo_obj.user_type if userinfo_obj.user_type else 'N/A',
        'user_name': userinfo_obj.username,
        'created_dt': userinfo_obj.created_dt,
        'year': userinfo_obj.year if userinfo_obj.year else -1,
        'school_id': userinfo_obj.school_id if userinfo_obj.school_id else -1,
        'school_name': userinfo_obj.school.school_name if userinfo_obj.school else 'N/A',
        'student_grade': userinfo_obj.student_grade if userinfo_obj.student_grade else -1,
        'student_class': userinfo_obj.student_class if userinfo_obj.student_class else -1,
        'student_number': userinfo_obj.student_number if userinfo_obj.student_number else -1,
        'student_name': userinfo_obj.student_name if userinfo_obj.student_name else 'N/A',
        'phone_number': userinfo_obj.phone_number if userinfo_obj.phone_number else 'N/A',
        'user_display_name': userinfo_obj.user_display_name if userinfo_obj.user_display_name else 'N/A',
        'dob': userinfo_obj.dob if userinfo_obj.dob else 'N/A',
        'gender': userinfo_obj.gender if userinfo_obj.gender else 'N/A',
        'height': userinfo_obj.height if userinfo_obj.height else -1,
        'dob': userinfo_obj.dob if userinfo_obj.dob else 'N/A',
    }


def fetch_recent_mails(email_host, email_user, email_password, minutes=1,
                       allowed_domains=["vmms.nate.com", "ktfmms.magicn.com", "mmsmail.uplus.co.kr", "lguplus.com"]):
    # Connect to the email server
    mail = imaplib.IMAP4_SSL(email_host)
    mail.login(email_user, email_password)

    mail.select('inbox')

    # Calculate the time window for emails
    kst = timezone(timedelta(hours=9))  # Korean Standard Time (UTC+9)
    end_time = datetime.now(kst)
    start_time = end_time - timedelta(minutes=minutes)

    # Format times for IMAP search query
    since_time = start_time.strftime('%d-%b-%Y %H:%M:%S')

    fetched_data = defaultdict(list)

    # Search for emails received within the last 'minutes' minutes
    result, data = mail.uid('search', None, f'(SINCE "{since_time}")')
    if result == 'OK':
        all_email = data[0].split()
        for uid in all_email:
            result, msg_data = mail.uid('fetch', uid, '(RFC822)')

            if result == 'OK':
                msg = email.message_from_bytes(msg_data[0][1], policy=default)

                # Extract the email's date
                email_date = email.utils.parsedate_to_datetime(msg['Date'])
                if email_date is None:
                    continue

                # Ensure the email's date is within the last minute
                if not (start_time <= email_date <= end_time):
                    continue

                sender = msg['From'].replace('<', '').replace('>', '').replace('"', '')
                sender_domain = sender.split('@')[-1].strip()

                # Check if the sender's domain is in the allowed list
                if sender_domain in allowed_domains:
                    # Extract the subject of the email
                    mobile_uid = str(msg['Subject'])
                    phone_number = str(sender.split(' ')[0])
                    fetched_data['uid'].append(mobile_uid)
                    fetched_data['phone_number'].append(phone_number)

    # Cleanup
    mail.logout()

    return fetched_data


def extract_digits(text):
    return re.search(r'\d+', text).group()


""" CodeInfo는 캐싱처리 """
from django.core.cache import cache
from functools import lru_cache


### CodeInfo (DB의 BodyResult 값에 대한 정보) 반환 함수
@lru_cache(maxsize=None)
def get_code_info_dict():
    code_infos = CodeInfo.objects.all()
    return {
        code_info.code_id: {
            'min': code_info.normal_min_value,
            'max': code_info.normal_max_value,
            'code_name': code_info.code_name
        }
        for code_info in code_infos
    }


### 정상범위 비율 계산 함수
def calculate_normal_ratio(body_result):
    fields_to_check = [
        'face_level_angle',
        'shoulder_level_angle',
        'hip_level_angle',
        'leg_length_ratio',
        'left_leg_alignment_angle',
        'right_leg_alignment_angle',
        'left_back_knee_angle',
        'right_back_knee_angle',
        'forward_head_angle',
        'scoliosis_shoulder_ratio',
        'scoliosis_hip_ratio'
    ]

    code_info_dict = get_code_info_dict()
    true_count = 0
    status_results = {}

    for field in fields_to_check:
        value = getattr(body_result, field)
        if value is not None and field in code_info_dict:
            normal_range = code_info_dict[field]
            is_normal = normal_range['min'] <= value <= normal_range['max']
            if is_normal:
                true_count += 1
            # CodeInfo의 code_name을 키로 사용하여 상태 저장
            status_results[normal_range['code_name']] = '' if is_normal else '주의'

    return f"{true_count}/{len(fields_to_check)}", status_results


def create_excel_report(df, user_type, code_names):
    # 기존 엑셀 파일 생성 코드
    excel_buffer = excel_io.BytesIO()
    df.to_excel(excel_buffer, index=False, sheet_name='데이터')

    workbook = load_workbook(excel_buffer)
    worksheet = workbook['데이터']

    # 기존 스타일링 적용
    adjust_column_widths(worksheet)
    worksheet.auto_filter.ref = worksheet.dimensions
    apply_normal_range_highlighting(worksheet)

    # 통계 시트 추가
    add_summary_sheet(workbook, df, code_names)

    return workbook


# 엑셀 파일의 열 폭 조정 함수
def adjust_column_widths(worksheet):
    for column in worksheet.columns:  # '검사일' 열에 대한 폭 조정
        column_letter = get_column_letter(column[0].column)
        if worksheet[f"{column_letter}1"].value == "검사일":
            max_length = max(len(str(cell.value or "")) for cell in column)
            worksheet.column_dimensions[column_letter].width = max_length + 2

    # worksheet의 A 열이 '부서명' 이면 F-P 열까지의 폭을 15로 설정
    if worksheet['A1'].value == '부서명':  # 기관 조건
        for col_idx in range(ord('F') - ord('A') + 1, ord('P') - ord('A') + 2):
            column_letter = get_column_letter(col_idx)
            worksheet.column_dimensions[column_letter].width = 15
        return

        # 학교 조건
    for col_idx in range(ord('H') - ord('A') + 1, ord('R') - ord('A') + 2):
        column_letter = get_column_letter(col_idx)
        worksheet.column_dimensions[column_letter].width = 15


# 정상범위가 7/11 이하 일 경우 엑셀의 셀 색상 변경 함수
def apply_normal_range_highlighting(worksheet):
    highlight_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")

    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        for cell in row:
            if cell.value and '정상범위' in str(worksheet.cell(row=1, column=cell.column).value):
                try:
                    numerator, denominator = map(int, str(cell.value).split("/"))
                    if numerator / denominator <= 7 / 11:  # 7/11 이하일 경우 색상 변경 -> 주의를 시각적으로 나타내기 위함
                        cell.fill = highlight_fill
                except:
                    pass


# 통계 엑셀 시트 생성 함수
def add_summary_sheet(workbook, df, code_names):
    # 새로운 시트 생성
    summary_sheet = workbook.create_sheet("통계", 0)

    # 총 인원수와 검사 받은 인원수 계산
    total_count = len(df)
    examined_count = len(df[df['검사결과'] == 'O'])

    # 각 code별 주의 인원수 계산
    warning_counts = {}
    for code in code_names:
        warning_counts[code] = len(df[df[code] == '주의'])

    # 데이터 추가
    summary_sheet['A1'] = '항목'
    summary_sheet['B1'] = '인원수'

    summary_sheet['A2'] = '총 인원수'
    summary_sheet['B2'] = total_count

    summary_sheet['A3'] = '검사를 받은 인원수'
    summary_sheet['B3'] = examined_count

    # code별 주의 인원수 추가
    for idx, (code, count) in enumerate(warning_counts.items(), start=4):
        summary_sheet[f'A{idx}'] = f'{code} 주의 인원수'
        summary_sheet[f'B{idx}'] = count

    # 열 너비 조정
    summary_sheet.column_dimensions['A'].width = 25
    summary_sheet.column_dimensions['B'].width = 15